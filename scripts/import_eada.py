"""Pull athletics participation and spending from EADA into the sample database.

A second federal source, because the first one does not carry this. IPEDS has
no athletics participation and no athletics money at all — only membership
flags like `member_ncaa`. Everything in this area comes from EADA, the Equity
in Athletics Disclosure Act collection, published by the Department of
Education as one bulk zip per survey year.

    uv run python scripts/import_eada.py

Run it **after** `scripts/import_ipeds.py`, which drops and recreates
`ingest_runs`. This script deletes and re-inserts only its own rows there, so
running it twice is harmless, but running the IPEDS ingest afterwards will wipe
its year metadata. `app/db.py` falls back to the table's own years when that
happens, so the app degrades to "works but cannot say the series ended" rather
than breaking.

Do not use EADA's JSON API for history. It accepts a year parameter and then
ignores it: `?year=2020` and `?surveyYear=2020` both return the newest survey.
The bulk files below are the only route to prior years.

**Year alignment.** EADA labels a file by the academic year it ends —
`EADA_2024-2025.zip` is the 2024-25 year, reported July 2024 to June 2025. That
is the cohort IPEDS calls 2024, so the file's label is decremented on the way
in and every year in this table lines up with the IPEDS tables beside it.
"""

import argparse
import io
import sqlite3
import urllib.request
import zipfile
from concurrent.futures import ThreadPoolExecutor
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook
from schools import UNITIDS

FILE_URL = "https://ope.ed.gov/athletics/api/dataFiles/file?fileName=EADA_{a}-{b}.zip"
DB_PATH = Path(__file__).resolve().parent.parent / "data" / "likeforlike.db"
TABLE = "eada_institutions"

# EADA file labels. 2025 is the newest published; the next lands around spring
# 2027. Six years is enough for a trend without downloading 150MB.
FILE_YEARS = range(2020, 2026)

# Only the columns the athletics area reads. The file carries 168; keeping the
# rest would triple the table for data nothing displays.
#
# UNDUP_CT_* rather than PARTIC_*: the latter counts a cross-country runner who
# also runs track twice, which overstates the athlete share of a student body
# by a fifth at some schools.
COLUMNS = {
    "unitid": "INTEGER",
    "EFTotalCount": "INTEGER",
    "UNDUP_CT_PARTIC_MEN": "INTEGER",
    "UNDUP_CT_PARTIC_WOMEN": "INTEGER",
    "STUDENTAID_MEN": "INTEGER",
    "STUDENTAID_WOMEN": "INTEGER",
    "STUDENTAID_TOTAL": "INTEGER",
    "RECRUITEXP_TOTAL": "INTEGER",
    "GRND_TOTAL_EXPENSE": "INTEGER",
    "GRND_TOTAL_REVENUE": "INTEGER",
    "HDCOACH_SALARY_MEN": "INTEGER",
    "HDCOACH_SALARY_WOMEN": "INTEGER",
    "classification_name": "TEXT",
}


def fetch(file_year: int) -> tuple[int, list[dict], str]:
    """One survey year's institution-level rows for the sample schools."""
    url = FILE_URL.format(a=file_year - 1, b=file_year)
    with urllib.request.urlopen(url, timeout=300) as response:
        blob = response.read()

    with zipfile.ZipFile(io.BytesIO(blob)) as archive:
        name = next(n for n in archive.namelist() if n.lower().endswith("instlevel.xlsx"))
        with archive.open(name) as handle:
            workbook = load_workbook(io.BytesIO(handle.read()), read_only=True)

    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = [str(c) for c in next(rows)]
    lower = [c.lower() for c in header]
    index = {c: lower.index(c.lower()) for c in COLUMNS if c.lower() in lower}

    wanted = set(UNITIDS)
    out = []
    for row in rows:
        if row[index["unitid"]] in wanted:
            record = {c: row[i] for c, i in index.items()}
            # IPEDS calls the 2024-25 year 2024; EADA labels it 2025.
            record["year"] = file_year - 1
            out.append(record)
    return file_year, out, url


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--db", type=Path, default=DB_PATH)
    args = parser.parse_args()

    if not args.db.exists():
        raise SystemExit(f"No database at {args.db}. Run scripts/import_ipeds.py first.")

    connection = sqlite3.connect(args.db)

    print(f"Fetching {len(FILE_YEARS)} EADA survey years...\n")
    with ThreadPoolExecutor(max_workers=6) as pool:
        results = sorted(pool.map(fetch, FILE_YEARS))

    definition = ", ".join(f'"{c}" {t}' for c, t in COLUMNS.items())
    connection.execute(f'DROP TABLE IF EXISTS "{TABLE}"')
    connection.execute(f'CREATE TABLE "{TABLE}" ("year" INTEGER NOT NULL, {definition})')

    fields = ["year", *COLUMNS]
    placeholders = ",".join("?" * len(fields))
    connection.executemany(
        f'INSERT INTO "{TABLE}" VALUES ({placeholders})',
        [tuple(r.get(f) for f in fields) for _, rows, _ in results for r in rows],
    )

    # Only this table's rows, so a re-run is idempotent and the IPEDS ingest's
    # own metadata is left alone.
    connection.execute("DELETE FROM ingest_runs WHERE table_name = ?", (TABLE,))
    for file_year, rows, url in results:
        covered = len({r["unitid"] for r in rows})
        connection.execute(
            "INSERT INTO ingest_runs VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (
                TABLE,
                "athletics",
                file_year - 1,
                url,
                len(rows),
                covered,
                datetime.now(UTC).isoformat(),
                "If I play a sport, how big a part of this place is that?",
            ),
        )
        print(f"  EADA {file_year - 1}-{file_year}  ->  year {file_year - 1}"
              f"  {len(rows):>3} rows  {covered:>2}/25 schools")

    connection.commit()
    connection.close()
    print(f"\nWrote {TABLE} to {args.db}")


if __name__ == "__main__":
    main()
